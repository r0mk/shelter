#!/usr/bin/python3
#Install: pip3 install openpyxl 
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import datetime
from tkinter import filedialog
from tkinter import *
 
if len(sys.argv) < 2:
    print('Usage ./script file.xlsx')
    root = Tk()
    root.filename =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("Excel files","*.xlsx"),("all files","*.*")))
    print (root.filename)
    wb2 = load_workbook(root.filename,data_only = True)
    #sys.exit()
if not root.filename:
    try:
        wb2 = load_workbook(sys.argv[1],data_only = True)
    except:
        print('File {} not found'.format(sys.argv[1]))

sheet_ranges = wb2[wb2.sheetnames[0]]
rows = sheet_ranges.iter_rows(min_row=1, max_row=1)
employee_id = sheet_ranges.iter_cols(min_col=1, max_col=1)
not_needed = ['Totals', 'Labor Cost', 'Actual Sales', 'Labor %',  'Unpaid Breaks', 'Hourly Rate', 'Total', 'Employee ID', 'Position']

first_row = next(rows)
#row_lenth = len(first_row)
for cell in first_row:
    if cell.value and cell.value not in not_needed:
        #print('column name ' + str(cell.value))
        #print(dir(cell.value))
        pass
year = sheet_ranges['D1'].value.year
month = sheet_ranges['D1'].value.month

first_col = next(employee_id)
col_lenth = len(first_col)

with open('Output.csv','w') as f:
    f.write('Loonjaar,Periode,Medewerker,Type,Salaris component,Invoer type,Waarde\n')

for cell in first_col:
    earned = []
    if cell.value and cell.value not in not_needed:
        earned = ([sheet_ranges['AI'][cell.row - 1].value])
        for pos_counter in range(1,10):
            if sheet_ranges['B' + str(cell.row + pos_counter)].value == None and sheet_ranges['A' + str(cell.row + pos_counter)].value == None:
                earned.append(sheet_ranges['AI'][cell.row + pos_counter - 1].value)
                pos_counter = pos_counter + 1
            else:
                break
        worked_days = 0
        worked_hours = []
        for day in range(1,32):
            for position in range(0,pos_counter):
                if sheet_ranges[get_column_letter(day + 3)][cell.row + position -1].value:
                    worked_days = worked_days + 1
                    worked_hours.append(sheet_ranges[get_column_letter(day + 3 )][cell.row + position -1].value)
        try:
            print('{},{},{},1,,1,{}'.format(year,month,cell.value,worked_days))
            print('{},{},{},2,,1,{}'.format(year,month,cell.value,round(sum(worked_hours))))
            with open('Output.csv','a') as f:
                f.write('{},{},{},1,,1,{}\n'.format(year,month,cell.value,worked_days))
                f.write('{},{},{},2,,1,{}\n'.format(year,month,cell.value,round(sum(worked_hours))))
        except:
            pass
        #sys.exit()
